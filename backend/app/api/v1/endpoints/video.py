from typing import List
import json
import uuid as uuid_module
import asyncio
from datetime import datetime, timedelta, timezone

from typing import Optional
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, select
from loguru import logger

from app.schema.video_schema import (
    GenerateUploadURLRequest,
    GenerateUploadURLResponse,
    GenerateDownloadURLRequest,
    GenerateDownloadURLResponse,
    UploadCompleteRequest,
    UploadCompleteResponse,
    GenerateExcelUploadURLRequest,
    GenerateExcelUploadURLResponse,
    RenameVideoRequest,
    RenameVideoResponse,
    DeleteVideoResponse,
    VideoResponse,
    BatchUploadCompleteRequest,
    BatchUploadCompleteResponse,
    LiveCaptureRequest,
    LiveCaptureResponse,
    LiveCheckResponse,
)
from app.services.video_service import VideoService
from app.repository.video_repository import VideoRepository
from app.core.dependencies import get_db, get_current_user
from app.utils.video_progress import calculate_progress, get_status_message
from app.core.container import Container
from app.models.orm.upload import Upload
from app.models.orm.video import Video

router = APIRouter(
    prefix="/videos",
    tags=["videos"],
)

# Initialize service (could be injected via DI container)
video_service = VideoService()


def _replace_blob_url_to_cdn(url: str) -> str:
    """Replace blob storage domain with CDN domain if applicable."""
    if url and isinstance(url, str):
        return url.replace(
            "https://aitherhub.blob.core.windows.net",
            "https://cdn.aitherhub.com"
        )
    return url


@router.post("/generate-upload-url", response_model=GenerateUploadURLResponse)
async def generate_upload_url(
    payload: GenerateUploadURLRequest,
    db: AsyncSession = Depends(get_db),
):
    try:
        result = await video_service.generate_upload_url(
            email=payload.email,
            db=db,
            video_id=payload.video_id,
            filename=payload.filename,
        )
        return GenerateUploadURLResponse(**result)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to generate upload URL: {exc}")


@router.post("/generate-download-url", response_model=GenerateDownloadURLResponse)
async def generate_download_url(payload: GenerateDownloadURLRequest):
    try:
        result = await video_service.generate_download_url(
            email=payload.email,    
            video_id=payload.video_id,
            filename=payload.filename,
            expires_in_minutes=payload.expires_in_minutes,
        )
        return GenerateDownloadURLResponse(**result)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to generate download URL: {exc}")


@router.post("/upload-complete", response_model=UploadCompleteResponse)
async def upload_complete(
    payload: UploadCompleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Handle upload completion - save video info to database"""
    try:
        # Verify email matches current user (current_user is a dict)
        if current_user["email"] != payload.email:
            raise HTTPException(status_code=403, detail="Email does not match current user")
        
        # Initialize video service with repository
        video_repo = VideoRepository(lambda: db)
        service = VideoService(video_repository=video_repo)
        
        result = await service.handle_upload_complete(
            user_id=current_user["id"],
            email=payload.email,
            video_id=payload.video_id,
            original_filename=payload.filename,
            db=db,
            upload_id=payload.upload_id,
            upload_type=payload.upload_type or "screen_recording",
            excel_product_blob_url=payload.excel_product_blob_url,
            excel_trend_blob_url=payload.excel_trend_blob_url,
            time_offset_seconds=payload.time_offset_seconds or 0,
        )
        return UploadCompleteResponse(**result)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to complete upload: {exc}")


@router.post("/batch-upload-complete", response_model=BatchUploadCompleteResponse)
async def batch_upload_complete(
    payload: BatchUploadCompleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Handle batch upload completion - multiple videos sharing the same Excel files"""
    try:
        if current_user["email"] != payload.email:
            raise HTTPException(status_code=403, detail="Email does not match current user")

        video_repo = VideoRepository(lambda: db)
        service = VideoService(video_repository=video_repo)

        video_ids = []
        for v in payload.videos:
            result = await service.handle_upload_complete(
                user_id=current_user["id"],
                email=payload.email,
                video_id=v.video_id,
                original_filename=v.filename,
                db=db,
                upload_id=v.upload_id,
                upload_type="clean_video",
                excel_product_blob_url=payload.excel_product_blob_url,
                excel_trend_blob_url=payload.excel_trend_blob_url,
                time_offset_seconds=v.time_offset_seconds or 0,
            )
            video_ids.append(result["video_id"])

        return BatchUploadCompleteResponse(
            video_ids=video_ids,
            status="uploaded",
            message=f"{len(video_ids)} videos queued for analysis",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to complete batch upload: {exc}")


@router.post("/generate-excel-upload-url", response_model=GenerateExcelUploadURLResponse)
async def generate_excel_upload_url(
    payload: GenerateExcelUploadURLRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Generate SAS upload URLs for Excel files (product + trend_stats)"""
    try:
        service = VideoService()
        result = await service.generate_excel_upload_urls(
            email=payload.email,
            video_id=payload.video_id,
            product_filename=payload.product_filename,
            trend_filename=payload.trend_filename,
        )
        return GenerateExcelUploadURLResponse(**result)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel upload URLs: {exc}")


@router.get("/uploads/check/{user_id}")
async def check_upload_resume(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Check if user has an in-progress upload to resume.

    Returns upload_resume=True only when:
    - An Upload record exists for this user, AND
    - The record is less than 24 hours old, AND
    - No corresponding Video record exists that was created after the upload
      (which would indicate the upload already completed successfully)
    """
    try:
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        result = await db.execute(
            select(Upload)
            .where(Upload.user_id == user_id)
            .order_by(Upload.created_at.desc(), Upload.id.desc())
            .limit(1)
        )
        upload = result.scalar_one_or_none()

        if not upload:
            return {"upload_resume": False}

        # Check 1: If upload record is older than 24 hours, treat as stale
        now = datetime.now(timezone.utc)
        upload_created = (
            upload.created_at.replace(tzinfo=timezone.utc)
            if upload.created_at.tzinfo is None
            else upload.created_at
        )
        upload_age = now - upload_created
        if upload_age > timedelta(hours=24):
            logger.info(
                f"Stale upload record {upload.id} for user {user_id} "
                f"(age: {upload_age}). Deleting."
            )
            await db.delete(upload)
            await db.commit()
            return {"upload_resume": False}

        # Check 2: If any Video record exists for this user that was created
        # around the same time or after the Upload record, the upload has
        # already completed successfully — clean up and return false.
        video_result = await db.execute(
            select(Video)
            .where(
                Video.user_id == user_id,
                Video.status != "NEW",
            )
            .order_by(Video.created_at.desc())
            .limit(1)
        )
        latest_video = video_result.scalar_one_or_none()

        if latest_video and latest_video.created_at:
            video_created = (
                latest_video.created_at.replace(tzinfo=timezone.utc)
                if latest_video.created_at.tzinfo is None
                else latest_video.created_at
            )
            # If the latest video was created within 5 min before or after
            # the upload record, the upload is already complete
            if video_created >= upload_created - timedelta(minutes=5):
                logger.info(
                    f"Upload {upload.id} already completed "
                    f"(video {latest_video.id} status={latest_video.status}). "
                    f"Cleaning up stale upload record."
                )
                await db.delete(upload)
                await db.commit()
                return {"upload_resume": False}

        # Check 3: If any video for this user is currently being processed
        # (not in a terminal state), the upload was successful and processing
        # is underway — no need to show resume dialog.
        processing_result = await db.execute(
            select(Video)
            .where(
                Video.user_id == user_id,
                Video.status.notin_(["NEW", "DONE", "ERROR", "uploaded"]),
            )
            .limit(1)
        )
        processing_video = processing_result.scalar_one_or_none()
        if processing_video:
            logger.info(
                f"Upload {upload.id} has a video in processing "
                f"(video {processing_video.id} status={processing_video.status}). "
                f"Cleaning up upload record."
            )
            await db.delete(upload)
            await db.commit()
            return {"upload_resume": False}

        return {"upload_resume": True, "upload_id": str(upload.id)}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Failed to check upload resume for user {user_id}: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to check upload resume: {exc}")


@router.delete("/uploads/clear/{user_id}")
async def clear_user_uploads(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Clear all in-progress uploads for a user."""
    try:
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        # Delete all upload records for this user
        result = await db.execute(
            select(Upload).where(Upload.user_id == user_id)
        )
        uploads = result.scalars().all()
        deleted_count = len(uploads)

        for upload in uploads:
            await db.delete(upload)

        await db.commit()

        return {
            "status": "success",
            "message": f"Deleted {deleted_count} upload record(s) for user {user_id}",
            "deleted_count": deleted_count,
        }
    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to clear uploads: {exc}")



@router.get("/user/{user_id}", response_model=List[VideoResponse])
async def get_videos_by_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Return list of videos for the given `user_id`.

    This endpoint requires authentication and only allows a user to fetch their own videos.
    """
    try:
        # Enforce that a user can only access their own videos
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        video_repo = VideoRepository(lambda: db)
        videos = await video_repo.get_videos_by_user(user_id=user_id)

        return [VideoResponse.from_orm(v) for v in videos]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to fetch videos: {exc}")




@router.get("/user/{user_id}/with-clips")
async def get_videos_by_user_with_clips(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Return list of videos for the given `user_id` with clip counts.
    This is used by the sidebar to show clip availability indicators.
    """
    try:
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        # Get videos with clip counts + sales/duration summary + memo count in a single query
        sql = text("""
            SELECT v.id, v.original_filename, v.status,
                   v.upload_type, v.created_at, v.updated_at,
                   COALESCE(c.clip_count, 0) as clip_count,
                   COALESCE(c.completed_count, 0) as completed_clip_count,
                   p.total_gmv,
                   p.max_time_end,
                   COALESCE(m.memo_count, 0) as memo_count,
                   v.top_products as top_products_json
            FROM videos v
            LEFT JOIN (
                SELECT video_id,
                       COUNT(DISTINCT phase_index) as clip_count,
                       COUNT(DISTINCT CASE WHEN status = 'completed' THEN phase_index END) as completed_count
                FROM video_clips
                GROUP BY video_id
            ) c ON v.id = c.video_id
            LEFT JOIN (
                SELECT video_id,
                       SUM(COALESCE(gmv, 0)) as total_gmv,
                       MAX(time_end) as max_time_end
                FROM video_phases
                GROUP BY video_id
            ) p ON v.id = p.video_id
            LEFT JOIN (
                SELECT video_id,
                       COUNT(*) as memo_count
                FROM video_phases
                WHERE (user_comment IS NOT NULL AND user_comment != '')
                   OR (user_rating IS NOT NULL AND user_rating > 0)
                GROUP BY video_id
            ) m ON v.id = m.video_id
            WHERE v.user_id = :user_id
            ORDER BY v.created_at DESC
        """)
        result = await db.execute(sql, {"user_id": user_id})
        rows = result.fetchall()

        import json as _json

        videos = []
        for row in rows:
            vid = str(row.id)
            # Parse cached top_products from videos table
            top_prods = []
            if row.top_products_json:
                try:
                    top_prods = _json.loads(row.top_products_json)
                except (ValueError, TypeError):
                    top_prods = []
            videos.append({
                "id": vid,
                "original_filename": row.original_filename,
                "status": row.status,
                "upload_type": row.upload_type,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
                "clip_count": row.clip_count,
                "completed_clip_count": row.completed_clip_count,
                "total_gmv": float(row.total_gmv) if row.total_gmv and float(row.total_gmv) > 0 else None,
                "stream_duration": float(row.max_time_end) if row.max_time_end else None,
                "memo_count": row.memo_count,
                "top_products": top_prods,
            })

        return videos

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to fetch videos with clips: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch videos: {exc}")


@router.delete("/{video_id}", response_model=DeleteVideoResponse)
async def delete_video(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Delete a video and its related data (only owner can delete)."""
    try:
        user_id = current_user["id"]
        video_repo = VideoRepository(lambda: db)

        # Delete ALL related records first to avoid FK constraint violations
        # Order matters: delete child tables before parent tables
        # Use safe_delete to skip tables that may not exist yet
        tables_to_delete = [
            # Level 3: grandchild tables (FK to child tables)
            "DELETE FROM speech_segments WHERE audio_chunk_id IN (SELECT id FROM audio_chunks WHERE video_id = :vid)",
            "DELETE FROM frame_analysis_results WHERE frame_id IN (SELECT id FROM video_frames WHERE video_id = :vid)",
            # Level 2: child tables with video_id FK
            "DELETE FROM video_frames WHERE video_id = :vid",
            "DELETE FROM video_product_exposures WHERE video_id = :vid",
            "DELETE FROM video_clips WHERE video_id = :vid",
            "DELETE FROM audio_chunks WHERE video_id = :vid",
            "DELETE FROM chats WHERE video_id = :vid",
            "DELETE FROM group_best_phases WHERE video_id = :vid",
            "DELETE FROM phase_insights WHERE video_id = :vid",
            "DELETE FROM video_phases WHERE video_id = :vid",
            "DELETE FROM video_insights WHERE video_id = :vid",
            "DELETE FROM processing_jobs WHERE video_id = :vid",
            "DELETE FROM reports WHERE video_id = :vid",
            "DELETE FROM video_processing_state WHERE video_id = :vid",
            # Structure tables
            "DELETE FROM video_structure_group_best_videos WHERE video_id = :vid",
            "DELETE FROM video_structure_group_members WHERE video_id = :vid",
            "DELETE FROM video_structure_features WHERE video_id = :vid",
        ]

        for sql in tables_to_delete:
            try:
                await db.execute(text(sql), {"vid": video_id})
            except Exception as table_err:
                # Skip if table doesn't exist (e.g., migration not yet applied)
                logger.warning(f"Skipping delete for non-existent table: {table_err}")
                await db.rollback()
                # Re-start transaction for next delete
                continue

        await db.commit()

        # Delete the video record
        deleted = await video_repo.delete_video(video_id=video_id, user_id=user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Video not found or not owned by user")

        return DeleteVideoResponse(id=video_id, message="Video deleted successfully")
    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete video: {exc}")


@router.patch("/{video_id}/rename", response_model=RenameVideoResponse)
async def rename_video(
    video_id: str,
    payload: RenameVideoRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Rename a video (only owner can rename)."""
    try:
        user_id = current_user["id"]
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.rename_video(
            video_id=video_id, user_id=user_id, new_name=payload.name
        )
        if not video:
            raise HTTPException(status_code=404, detail="Video not found or not owned by user")

        return RenameVideoResponse(
            id=str(video.id),
            original_filename=video.original_filename,
            message="Video renamed successfully",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to rename video: {exc}")


@router.get("/{video_id}/status/stream")
async def stream_video_status(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Stream video processing status updates via Server-Sent Events (SSE).

    This endpoint provides real-time status updates for video processing.
    It polls the database every 2 seconds and sends status changes to the client.
    The stream automatically closes when processing reaches DONE or ERROR status.
    Supports long-running videos up to 4 hours with heartbeat messages every 30 seconds.

    Args:
        video_id: UUID of the video to monitor
        db: Database session
        current_user: Authenticated user

    Returns:
        StreamingResponse with SSE events containing:
        - status: Current processing status
        - progress: Progress percentage (0-100)
        - message: User-friendly Japanese status message
        - updated_at: Timestamp of last update
        - heartbeat: Boolean indicating heartbeat message (sent every 30 seconds)

    Example SSE events:
        data: {"video_id": "...", "status": "STEP_3_TRANSCRIBE_AUDIO", "progress": 40, "message": "音声書き起こし中...", "updated_at": "2026-01-20T..."}
        data: {"heartbeat": true, "timestamp": "2026-01-20T...", "poll_count": 15}
    """

    async def event_generator():
        last_status = None
        last_step_progress = None
        poll_count = 0
        max_polls = 7200  # 4 hours max for long videos (7200 * 2 seconds = 14400 seconds = 4 hours)

        try:
            # Verify video exists and ownership
            video_repo = VideoRepository(lambda: db)
            video = await video_repo.get_video_by_id(video_id)

            if not video:
                yield f"data: {json.dumps({'error': 'Video not found'})}\n\n"
                return

            if current_user and current_user.get("id") != video.user_id:
                yield f"data: {json.dumps({'error': 'Forbidden'})}\n\n"
                return

            # Stream status updates
            while poll_count < max_polls:
                try:
                    # Refresh video data
                    video = await video_repo.get_video_by_id(video_id)

                    if not video:
                        yield f"data: {json.dumps({'error': 'Video not found'})}\n\n"
                        break

                    current_status = video.status
                    current_step_progress = getattr(video, 'step_progress', None) or 0

                    # Send update if status changed OR step_progress changed
                    if current_status != last_status or current_step_progress != last_step_progress:
                        progress = calculate_progress(current_status)
                        message = get_status_message(current_status)

                        payload = {
                            "video_id": str(video.id),
                            "status": current_status,
                            "progress": progress,
                            "step_progress": current_step_progress,
                            "message": message,
                            "updated_at": video.updated_at.isoformat() if video.updated_at else None,
                        }

                        yield f"data: {json.dumps(payload)}\n\n"
                        last_status = current_status
                        last_step_progress = current_step_progress

                        logger.info(f"SSE: Video {video_id} status={current_status} step_progress={current_step_progress}%")

                    # Send heartbeat every 30 seconds (15 * 2 seconds) to keep connection alive
                    if poll_count > 0 and poll_count % 15 == 0:
                        heartbeat_payload = {
                            "heartbeat": True,
                            "timestamp": datetime.utcnow().isoformat(),
                            "poll_count": poll_count
                        }
                        yield f"data: {json.dumps(heartbeat_payload)}\n\n"
                        logger.debug(f"SSE: Heartbeat sent for video {video_id} (poll {poll_count})")

                    # Stop streaming if processing complete or error
                    if current_status in ["DONE", "ERROR"]:
                        yield "data: [DONE]\n\n"
                        logger.info(f"SSE: Video {video_id} processing completed with status {current_status}")
                        break

                    # Poll every 2 seconds
                    await asyncio.sleep(2)
                    poll_count += 1

                except Exception as e:
                    logger.error(f"SSE poll error for video {video_id}: {e}")
                    yield f"data: {json.dumps({'error': f'Poll error: {str(e)}'})}\n\n"
                    break

            # Timeout reached
            if poll_count >= max_polls:
                logger.warning(f"SSE: Video {video_id} stream timeout after {max_polls * 2} seconds")
                yield f"data: {json.dumps({'error': 'Stream timeout'})}\n\n"

        except Exception as e:
            logger.error(f"SSE stream error for video {video_id}: {e}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",  # Disable nginx buffering
            "Connection": "keep-alive",
        }
    )


@router.get("/{video_id}")
async def get_video_detail(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
        Video detail endpoint returning report 1 data.
        Optimized: single combined query, inline SAS generation, no ORM overhead.
    """
    import time as _time
    import os as _os
    from azure.storage.blob import generate_blob_sas as _generate_blob_sas, BlobSasPermissions as _BlobSasPermissions

    try:
        _t0 = _time.monotonic()

        # ---- Step 1: Single query to get video + user email ----
        sql_video = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.upload_type, v.excel_product_blob_url, v.excel_trend_blob_url,
                   v.compressed_blob_url,
                   u.email
            FROM videos v
            JOIN users u ON v.user_id = u.id
            WHERE v.id = :video_id
        """)
        vres = await db.execute(sql_video, {"video_id": video_id})
        video_row = vres.fetchone()
        if not video_row:
            raise HTTPException(status_code=404, detail="Video not found")

        if current_user and current_user.get("id") != video_row.user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        email = video_row.email
        compressed_blob = video_row.compressed_blob_url
        _t1 = _time.monotonic()

        # ---- Step 2: Parallel fetch phase_insights + video_phases + video_insights ----
        sql_combined = text("""
            SELECT
                vp.id as phase_id, vp.phase_index, vp.phase_description,
                vp.time_start, vp.time_end,
                COALESCE(vp.gmv, 0) as gmv,
                COALESCE(vp.order_count, 0) as order_count,
                COALESCE(vp.viewer_count, 0) as viewer_count,
                COALESCE(vp.like_count, 0) as like_count,
                COALESCE(vp.comment_count, 0) as comment_count,
                COALESCE(vp.share_count, 0) as share_count,
                COALESCE(vp.new_followers, 0) as new_followers,
                COALESCE(vp.product_clicks, 0) as product_clicks,
                COALESCE(vp.conversion_rate, 0) as conversion_rate,
                COALESCE(vp.gpm, 0) as gpm,
                COALESCE(vp.importance_score, 0) as importance_score,
                vp.product_names,
                vp.user_rating,
                vp.user_comment,
                vp.sas_token,
                vp.sas_expireddate,
                vp.cta_score,
                vp.audio_features,
                pi.insight
            FROM video_phases vp
            LEFT JOIN phase_insights pi ON pi.video_id = vp.video_id AND pi.phase_index = vp.phase_index
            WHERE vp.video_id = :video_id
            ORDER BY vp.phase_index ASC
        """)

        sql_latest_insight = text("""
            SELECT title, content
            FROM video_insights
            WHERE video_id = :video_id
            ORDER BY created_at DESC
            LIMIT 1
        """)

        # Execute both queries concurrently
        # Fallback: if cta_score/audio_features columns don't exist yet, retry without them
        has_cta_columns = True
        try:
            combined_task = db.execute(sql_combined, {"video_id": video_id})
            insight_task = db.execute(sql_latest_insight, {"video_id": video_id})
            combined_res, insight_res = await asyncio.gather(combined_task, insight_task)
        except Exception:
            has_cta_columns = False
            await db.rollback()
            sql_combined_fallback = text("""
                SELECT
                    vp.id as phase_id, vp.phase_index, vp.phase_description,
                    vp.time_start, vp.time_end,
                    COALESCE(vp.gmv, 0) as gmv,
                    COALESCE(vp.order_count, 0) as order_count,
                    COALESCE(vp.viewer_count, 0) as viewer_count,
                    COALESCE(vp.like_count, 0) as like_count,
                    COALESCE(vp.comment_count, 0) as comment_count,
                    COALESCE(vp.share_count, 0) as share_count,
                    COALESCE(vp.new_followers, 0) as new_followers,
                    COALESCE(vp.product_clicks, 0) as product_clicks,
                    COALESCE(vp.conversion_rate, 0) as conversion_rate,
                    COALESCE(vp.gpm, 0) as gpm,
                    COALESCE(vp.importance_score, 0) as importance_score,
                    vp.product_names,
                    vp.user_rating,
                    vp.user_comment,
                    vp.sas_token,
                    vp.sas_expireddate,
                    NULL as cta_score,
                    NULL as audio_features,
                    pi.insight
                FROM video_phases vp
                LEFT JOIN phase_insights pi ON pi.video_id = vp.video_id AND pi.phase_index = vp.phase_index
                WHERE vp.video_id = :video_id
                ORDER BY vp.phase_index ASC
            """)
            combined_task = db.execute(sql_combined_fallback, {"video_id": video_id})
            insight_task = db.execute(sql_latest_insight, {"video_id": video_id})
            combined_res, insight_res = await asyncio.gather(combined_task, insight_task)

        combined_rows = combined_res.fetchall()
        latest_insight = insight_res.fetchone()
        _t2 = _time.monotonic()

        # ---- Step 3: Build SAS URLs inline (no async service call needed) ----
        conn_str = _os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
        account_name = _os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "")
        container_name = _os.getenv("AZURE_BLOB_CONTAINER", "videos")
        account_key = ""
        for part in conn_str.split(";"):
            if part.startswith("AccountKey="):
                account_key = part.split("=", 1)[1]
                break

        now_utc = datetime.now(timezone.utc)
        now_naive = datetime.utcnow()
        sas_expiry = now_utc + timedelta(days=7)
        phases_needing_sas_update = []  # (phase_id, sas_url, expiry)

        def _make_sas_url(blob_name: str) -> str:
            """Generate SAS URL locally without any async/HTTP call."""
            sas = _generate_blob_sas(
                account_name=account_name,
                container_name=container_name,
                blob_name=blob_name,
                account_key=account_key,
                permission=_BlobSasPermissions(read=True),
                expiry=sas_expiry,
            )
            url = f"https://{account_name}.blob.core.windows.net/{container_name}/{blob_name}?{sas}"
            return _replace_blob_url_to_cdn(url)

        report1_items = []
        for r in combined_rows:
            # Check cached SAS
            video_clip_url = None
            if email and r.time_start is not None and r.time_end is not None:
                sas_token = r.sas_token
                sas_expire = r.sas_expireddate
                cache_valid = False
                if sas_token and sas_expire:
                    try:
                        if sas_expire.tzinfo is not None and sas_expire.tzinfo.utcoffset(sas_expire) is not None:
                            cache_valid = sas_expire.astimezone(timezone.utc) >= now_utc
                        else:
                            cache_valid = sas_expire >= now_naive
                    except Exception:
                        pass

                if cache_valid:
                    video_clip_url = sas_token
                elif account_key:
                    try:
                        ts = float(r.time_start)
                        te = float(r.time_end)
                        fname = f"{ts:.1f}_{te:.1f}.mp4"
                        blob_name = f"{email}/{video_id}/reportvideo/{fname}"
                        video_clip_url = _make_sas_url(blob_name)
                        if r.phase_id:
                            phases_needing_sas_update.append((r.phase_id, video_clip_url, sas_expiry))
                    except Exception:
                        video_clip_url = None

            # Parse product_names
            product_names_list = []
            pn_raw = r.product_names
            if pn_raw:
                try:
                    product_names_list = json.loads(pn_raw) if isinstance(pn_raw, str) else pn_raw
                except (json.JSONDecodeError, TypeError):
                    product_names_list = []

            # Only include phases that have insights (matching original behavior)
            if r.insight is not None:
                # Parse audio_features JSON text
                audio_features_parsed = None
                try:
                    if r.audio_features:
                        audio_features_parsed = json.loads(r.audio_features) if isinstance(r.audio_features, str) else r.audio_features
                except (json.JSONDecodeError, TypeError):
                    pass

                report1_items.append({
                    "phase_index": int(r.phase_index),
                    "phase_description": r.phase_description,
                    "time_start": r.time_start,
                    "time_end": r.time_end,
                    "insight": r.insight,
                    "video_clip_url": video_clip_url,
                    "user_rating": r.user_rating,
                    "user_comment": r.user_comment,
                    "cta_score": getattr(r, 'cta_score', None),
                    "audio_features": audio_features_parsed,
                    "csv_metrics": {
                        "gmv": r.gmv,
                        "order_count": r.order_count,
                        "viewer_count": r.viewer_count,
                        "like_count": r.like_count,
                        "comment_count": r.comment_count,
                        "share_count": r.share_count,
                        "new_followers": r.new_followers,
                        "product_clicks": r.product_clicks,
                        "conversion_rate": r.conversion_rate,
                        "gpm": r.gpm,
                        "importance_score": r.importance_score,
                        "product_names": product_names_list,
                    },
                })

        _t3 = _time.monotonic()

        # ---- Step 4: Batch persist new SAS tokens (fire-and-forget style) ----
        if phases_needing_sas_update:
            try:
                for pid, sas_url, exp_at in phases_needing_sas_update:
                    await db.execute(
                        text("UPDATE video_phases SET sas_token = :sas, sas_expireddate = :exp WHERE id = :id"),
                        {"sas": sas_url, "exp": exp_at, "id": pid}
                    )
                await db.commit()
            except Exception:
                pass  # Non-critical

        # ---- Step 5: Build report3 ----
        report3 = []
        if latest_insight:
            parsed = latest_insight.content
            try:
                if isinstance(parsed, str):
                    s = parsed.lstrip()
                    if s.startswith("{") or s.startswith("["):
                        parsed = json.loads(parsed)

                if isinstance(parsed, dict) and parsed.get("video_insights") and isinstance(parsed.get("video_insights"), list):
                    for item in parsed.get("video_insights"):
                        report3.append({"title": item.get("title"), "content": item.get("content")})
                elif isinstance(parsed, list):
                    for item in parsed:
                        report3.append({"title": item.get("title"), "content": item.get("content")})
                else:
                    report3.append({"title": latest_insight.title, "content": latest_insight.content})
            except Exception:
                report3.append({"title": latest_insight.title, "content": latest_insight.content})

        # ---- Step 6: Generate preview URL (inline, no service call) ----
        preview_url = None
        if compressed_blob and email and account_key:
            try:
                preview_filename = compressed_blob.split('/')[-1] if '/' in compressed_blob else compressed_blob
                blob_name = f"{email}/{video_id}/{preview_filename}"
                preview_url = _make_sas_url(blob_name)
            except Exception:
                preview_url = None

        _t_end = _time.monotonic()
        _perf = {
            "video_query_ms": round((_t1-_t0)*1000),
            "combined_query_ms": round((_t2-_t1)*1000),
            "build_response_ms": round((_t3-_t2)*1000),
            "total_ms": round((_t_end-_t0)*1000),
            "phase_count": len(combined_rows),
            "sas_generated": len(phases_needing_sas_update),
        }
        logger.info(f"[PERF] {_perf}")

        return {
            "id": str(video_row.id),
            "original_filename": video_row.original_filename,
            "status": video_row.status,
            "step_progress": getattr(video_row, 'step_progress', None) or 0,
            "upload_type": video_row.upload_type,
            "excel_product_blob_url": video_row.excel_product_blob_url,
            "excel_trend_blob_url": video_row.excel_trend_blob_url,
            "compressed_blob_url": compressed_blob,
            "preview_url": preview_url,
            "reports_1": report1_items,
            "report3": report3,
            "_perf": _perf,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to fetch video detail: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch video detail: {exc}")


@router.get("/{video_id}/product-data")
async def get_video_product_data(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Fetch and parse the product Excel file for a video.
    Returns parsed product data as JSON.
    Uses SAS tokens to access Azure Blob Storage (public access is disabled).
    """
    try:
        import httpx
        import tempfile
        import os
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions
        from datetime import timedelta

        # Get video's excel_product_blob_url and user email
        result = await db.execute(
            text("""
                SELECT v.excel_product_blob_url, v.excel_trend_blob_url, u.email
                FROM videos v
                JOIN users u ON v.user_id = u.id
                WHERE v.id = :vid
            """),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")

        product_blob_url = row[0]
        trend_blob_url = row[1]
        email = row[2]

        response_data = {
            "products": [],
            "trends": [],
            "has_product_data": False,
            "has_trend_data": False,
        }

        # Helper: generate SAS download URL from blob URL
        def _generate_sas_url(blob_url: str) -> str:
            """Generate a SAS-signed download URL from a raw blob URL."""
            conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
            account_name = ""
            account_key = ""
            for part in conn_str.split(";"):
                if part.startswith("AccountName="):
                    account_name = part.split("=", 1)[1]
                elif part.startswith("AccountKey="):
                    account_key = part.split("=", 1)[1]

            # Extract blob path from URL
            # URL format: https://account.blob.core.windows.net/videos/email/video_id/excel/filename.xlsx
            try:
                from urllib.parse import urlparse, unquote
                parsed = urlparse(blob_url)
                path = unquote(parsed.path)  # /videos/email/video_id/excel/filename.xlsx
                # Remove leading /videos/ to get blob_name
                if path.startswith("/videos/"):
                    blob_name = path[len("/videos/"):]
                else:
                    blob_name = path.lstrip("/")
                    # Remove container name if present
                    if blob_name.startswith("videos/"):
                        blob_name = blob_name[len("videos/"):]
            except Exception:
                # Fallback: construct blob name from email/video_id
                filename = blob_url.split("/")[-1].split("?")[0]
                blob_name = f"{email}/{video_id}/excel/{filename}"

            expiry = datetime.now(timezone.utc) + timedelta(minutes=30)
            sas = generate_blob_sas(
                account_name=account_name,
                container_name="videos",
                blob_name=blob_name,
                account_key=account_key,
                permission=BlobSasPermissions(read=True),
                expiry=expiry,
            )
            return f"https://{account_name}.blob.core.windows.net/videos/{blob_name}?{sas}"

        # Helper: download and parse Excel file
        async def _parse_excel(blob_url: str) -> list:
            """Download Excel via SAS URL and parse rows into list of dicts."""
            sas_url = _generate_sas_url(blob_url)
            import openpyxl
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(sas_url)
                if resp.status_code != 200:
                    logger.warning(f"Failed to download Excel (HTTP {resp.status_code}): {sas_url[:100]}...")
                    return []

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                    f.write(resp.content)
                    tmp_path = f.name

                try:
                    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                    ws = wb.active
                    items = []
                    if ws:
                        rows_data = list(ws.iter_rows(values_only=True))
                        if len(rows_data) >= 2:
                            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                            for data_row in rows_data[1:]:
                                if all(v is None for v in data_row):
                                    continue
                                item = {}
                                for i, val in enumerate(data_row):
                                    if i < len(headers):
                                        if val is None:
                                            item[headers[i]] = None
                                        elif isinstance(val, (int, float)):
                                            item[headers[i]] = val
                                        else:
                                            item[headers[i]] = str(val)
                                items.append(item)
                    wb.close()
                    return items
                finally:
                    os.unlink(tmp_path)

        # Parse product Excel
        if product_blob_url:
            try:
                products = await _parse_excel(product_blob_url)
                response_data["products"] = products
                response_data["has_product_data"] = len(products) > 0

                # Cache top 2 products by GMV in videos table
                if products:
                    try:
                        # Detect GMV and name columns
                        gmv_key = None
                        name_key = None
                        sample = products[0]
                        for k in sample.keys():
                            kl = k.lower() if k else ""
                            if "gmv" in kl:
                                gmv_key = k
                            if "商品名" in k or "product" in kl or "name" in kl:
                                name_key = k
                        if gmv_key and name_key:
                            sorted_products = sorted(
                                products,
                                key=lambda x: float(x.get(gmv_key, 0) or 0),
                                reverse=True,
                            )
                            top2 = []
                            for p in sorted_products[:2]:
                                pname = p.get(name_key, "")
                                if pname:
                                    # Truncate long product names
                                    pname = str(pname)[:50]
                                    top2.append(pname)
                            if top2:
                                import json as _json
                                await db.execute(
                                    text("UPDATE videos SET top_products = :tp WHERE id = :vid"),
                                    {"tp": _json.dumps(top2, ensure_ascii=False), "vid": video_id},
                                )
                                await db.commit()
                                logger.info(f"Cached top_products for video {video_id}: {top2}")
                    except Exception as cache_err:
                        logger.warning(f"Failed to cache top_products: {cache_err}")
            except Exception as e:
                logger.warning(f"Failed to parse product Excel: {e}")

        # Parse trend Excel
        if trend_blob_url:
            try:
                trends = await _parse_excel(trend_blob_url)
                response_data["trends"] = trends
                response_data["has_trend_data"] = len(trends) > 0
            except Exception as e:
                logger.warning(f"Failed to parse trend Excel: {e}")

        return response_data

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Failed to fetch product data: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch product data: {exc}")



# =========================
# Clip generation endpoints
# =========================

@router.post("/{video_id}/clips")
async def request_clip_generation(
    video_id: str,
    request_body: dict,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Request TikTok-style clip generation for a specific phase.
    
    Body:
    {
        "phase_index": 0,
        "time_start": 0.0,
        "time_end": 51.0,
        "speed_factor": 1.2  // optional, default 1.0 (1.0-1.5x)
    }
    """
    try:
        user_id = user.get("user_id") or user.get("id")
        phase_index = request_body.get("phase_index")
        time_start = request_body.get("time_start")
        time_end = request_body.get("time_end")
        speed_factor = float(request_body.get("speed_factor", 1.0))

        if phase_index is None or time_start is None or time_end is None:
            raise HTTPException(status_code=400, detail="phase_index, time_start, time_end are required")

        # Clamp speed_factor to safe range
        speed_factor = max(0.5, min(2.0, speed_factor))

        time_start = float(time_start)
        time_end = float(time_end)

        if time_end <= time_start:
            raise HTTPException(status_code=400, detail="time_end must be greater than time_start")

        # Check if clip already exists for this phase
        existing_sql = text("""
            SELECT id, status, clip_url
            FROM video_clips
            WHERE video_id = :video_id AND phase_index = :phase_index
            ORDER BY created_at DESC
            LIMIT 1
        """)
        existing = await db.execute(existing_sql, {"video_id": video_id, "phase_index": phase_index})
        existing_row = existing.fetchone()

        if existing_row:
            if existing_row.status == "completed" and existing_row.clip_url:
                # Already generated - return existing
                return {
                    "clip_id": str(existing_row.id),
                    "status": "completed",
                    "clip_url": _replace_blob_url_to_cdn(existing_row.clip_url),
                    "message": "Clip already generated",
                }
            elif existing_row.status in ("pending", "processing"):
                # Already in progress
                return {
                    "clip_id": str(existing_row.id),
                    "status": existing_row.status,
                    "message": "Clip generation already in progress",
                }
            # If failed, create a new one

        # Verify video belongs to user
        video_sql = text("SELECT id, user_id, original_filename FROM videos WHERE id = :video_id")
        vres = await db.execute(video_sql, {"video_id": video_id})
        video_row = vres.fetchone()

        if not video_row:
            raise HTTPException(status_code=404, detail="Video not found")
        if video_row.user_id != user_id:
            raise HTTPException(status_code=403, detail="Not authorized")

        # Get user email for blob path
        user_sql = text("SELECT email FROM users WHERE id = :user_id")
        ures = await db.execute(user_sql, {"user_id": user_id})
        user_row = ures.fetchone()
        email = user_row.email if user_row else None

        if not email:
            raise HTTPException(status_code=400, detail="User email not found")

        # Generate download SAS URL for source video
        from app.services.storage_service import generate_download_sas
        download_url, _ = await generate_download_sas(
            email=email,
            video_id=video_id,
            filename=video_row.original_filename,
            expires_in_minutes=1440,
        )

        # Create clip record
        clip_id = str(uuid_module.uuid4())
        insert_sql = text("""
            INSERT INTO video_clips (id, video_id, user_id, phase_index, time_start, time_end, status)
            VALUES (:id, :video_id, :user_id, :phase_index, :time_start, :time_end, 'pending')
        """)
        await db.execute(insert_sql, {
            "id": clip_id,
            "video_id": video_id,
            "user_id": user_id,
            "phase_index": phase_index,
            "time_start": time_start,
            "time_end": time_end,
        })
        await db.commit()

        # Enqueue clip generation job
        from app.services.queue_service import enqueue_job
        await enqueue_job({
            "job_type": "generate_clip",
            "clip_id": clip_id,
            "video_id": video_id,
            "blob_url": download_url,
            "time_start": time_start,
            "time_end": time_end,
            "phase_index": phase_index,
            "speed_factor": speed_factor,
        })

        logger.info(f"Clip generation requested: clip_id={clip_id}, video_id={video_id}, phase={phase_index}")

        return {
            "clip_id": clip_id,
            "status": "pending",
            "message": "Clip generation started",
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to request clip generation: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to request clip generation: {exc}")


@router.get("/{video_id}/clips/{phase_index}")
async def get_clip_status(
    video_id: str,
    phase_index: int,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Get clip generation status and download URL for a specific phase."""
    try:
        user_id = user.get("user_id") or user.get("id")

        sql = text("""
            SELECT id, status, clip_url, sas_token, sas_expireddate, error_message, created_at
            FROM video_clips
            WHERE video_id = :video_id AND phase_index = :phase_index
            ORDER BY created_at DESC
            LIMIT 1
        """)
        result = await db.execute(sql, {"video_id": video_id, "phase_index": phase_index})
        row = result.fetchone()

        if not row:
            return {
                "status": "not_found",
                "message": "No clip found for this phase",
            }

        response = {
            "clip_id": str(row.id),
            "status": row.status,
        }

        if row.status == "completed" and row.clip_url:
            # Generate or reuse SAS download URL
            clip_download_url = None

            # Check if existing SAS is still valid
            if row.sas_token and row.sas_expireddate:
                now = datetime.now(timezone.utc)
                expiry = row.sas_expireddate
                if expiry.tzinfo is None:
                    expiry = expiry.replace(tzinfo=timezone.utc)
                if expiry > now:
                    clip_download_url = row.sas_token

            if not clip_download_url:
                # Generate new SAS URL for clip
                try:
                    # Get user email
                    user_sql = text("SELECT email FROM users WHERE id = :user_id")
                    ures = await db.execute(user_sql, {"user_id": user_id})
                    user_row = ures.fetchone()

                    if user_row:
                        # Extract blob path from clip_url
                        from app.services.storage_service import generate_download_sas
                        # Parse the clip blob name from the URL
                        clip_url = row.clip_url
                        # clip_url format: https://account.blob.core.windows.net/container/email/video_id/clips/clip_X_Y.mp4
                        parts = clip_url.split("/")
                        # Find the email/video_id/clips/filename part
                        try:
                            container_idx = parts.index("videos") if "videos" in parts else -1
                            if container_idx >= 0 and container_idx + 1 < len(parts):
                                blob_path = "/".join(parts[container_idx + 1:])
                                from azure.storage.blob import generate_blob_sas, BlobSasPermissions
                                import os as _os
                                conn_str = _os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
                                account_name = ""
                                account_key = ""
                                for p in conn_str.split(";"):
                                    if p.startswith("AccountName="):
                                        account_name = p.split("=", 1)[1]
                                    if p.startswith("AccountKey="):
                                        account_key = p.split("=", 1)[1]

                                if account_name and account_key:
                                    expiry_dt = datetime.now(timezone.utc) + timedelta(hours=24)
                                    sas = generate_blob_sas(
                                        account_name=account_name,
                                        container_name="videos",
                                        blob_name=blob_path,
                                        account_key=account_key,
                                        permission=BlobSasPermissions(read=True),
                                        expiry=expiry_dt,
                                    )
                                    clip_download_url = f"https://{account_name}.blob.core.windows.net/videos/{blob_path}?{sas}"
                                    clip_download_url = _replace_blob_url_to_cdn(clip_download_url)

                                    # Cache the SAS token
                                    update_sql = text("""
                                        UPDATE video_clips
                                        SET sas_token = :sas_token, sas_expireddate = :expiry
                                        WHERE id = :id
                                    """)
                                    await db.execute(update_sql, {
                                        "sas_token": clip_download_url,
                                        "expiry": expiry_dt,
                                        "id": row.id,
                                    })
                                    await db.commit()
                        except Exception as e:
                            logger.warning(f"Failed to parse clip blob path: {e}")
                except Exception as e:
                    logger.warning(f"Failed to generate clip SAS: {e}")

            response["clip_url"] = clip_download_url or _replace_blob_url_to_cdn(row.clip_url)

        elif row.status == "failed":
            response["error_message"] = row.error_message

        return response

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to get clip status: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to get clip status: {exc}")


@router.get("/{video_id}/clips")
async def list_clips(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """List all clips for a video."""
    try:
        user_id = user.get("user_id") or user.get("id")

        sql = text("""
            SELECT id, phase_index, time_start, time_end, status, clip_url, sas_token, sas_expireddate, created_at
            FROM video_clips
            WHERE video_id = :video_id
            ORDER BY phase_index ASC, created_at DESC
        """)
        result = await db.execute(sql, {"video_id": video_id})
        rows = result.fetchall()

        clips = []
        seen_phases = set()
        for row in rows:
            # Only include the latest clip per phase
            if row.phase_index in seen_phases:
                continue
            seen_phases.add(row.phase_index)

            clip = {
                "clip_id": str(row.id),
                "phase_index": row.phase_index,
                "time_start": row.time_start,
                "time_end": row.time_end,
                "status": row.status,
            }
            if row.status == "completed" and row.clip_url:
                # Generate or reuse SAS download URL (same logic as get_clip_status)
                clip_download_url = None

                # Check if existing SAS is still valid
                if row.sas_token and row.sas_expireddate:
                    now = datetime.now(timezone.utc)
                    expiry = row.sas_expireddate
                    if expiry.tzinfo is None:
                        expiry = expiry.replace(tzinfo=timezone.utc)
                    if expiry > now:
                        clip_download_url = row.sas_token

                if not clip_download_url:
                    # Generate new SAS URL for clip
                    try:
                        clip_url = row.clip_url
                        parts = clip_url.split("/")
                        container_idx = parts.index("videos") if "videos" in parts else -1
                        if container_idx >= 0 and container_idx + 1 < len(parts):
                            blob_path = "/".join(parts[container_idx + 1:])
                            from azure.storage.blob import generate_blob_sas, BlobSasPermissions
                            import os as _os
                            conn_str = _os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
                            account_name = ""
                            account_key = ""
                            for p in conn_str.split(";"):
                                if p.startswith("AccountName="):
                                    account_name = p.split("=", 1)[1]
                                if p.startswith("AccountKey="):
                                    account_key = p.split("=", 1)[1]

                            if account_name and account_key:
                                expiry_dt = datetime.now(timezone.utc) + timedelta(hours=24)
                                sas = generate_blob_sas(
                                    account_name=account_name,
                                    container_name="videos",
                                    blob_name=blob_path,
                                    account_key=account_key,
                                    permission=BlobSasPermissions(read=True),
                                    expiry=expiry_dt,
                                )
                                clip_download_url = f"https://{account_name}.blob.core.windows.net/videos/{blob_path}?{sas}"
                                clip_download_url = _replace_blob_url_to_cdn(clip_download_url)

                                # Cache the SAS token
                                update_sql = text("""
                                    UPDATE video_clips
                                    SET sas_token = :sas_token, sas_expireddate = :expiry
                                    WHERE id = :id
                                """)
                                await db.execute(update_sql, {
                                    "sas_token": clip_download_url,
                                    "expiry": expiry_dt,
                                    "id": row.id,
                                })
                                await db.commit()
                    except Exception as e:
                        logger.warning(f"Failed to generate clip SAS in list: {e}")

                clip["clip_url"] = clip_download_url or _replace_blob_url_to_cdn(row.clip_url)
            clips.append(clip)

        return {"clips": clips}

    except Exception as exc:
        logger.exception(f"Failed to list clips: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to list clips: {exc}")



# ──────────────────────────────────────────────────────────────
# Phase Rating (Human Feedback)
# ──────────────────────────────────────────────────────────────

@router.put("/{video_id}/phases/{phase_index}/rating")
async def rate_phase(
    video_id: str,
    phase_index: int,
    request_body: dict,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Save a human rating (1-5) and optional comment for a specific phase.
    Also updates the quality_score in Qdrant for RAG learning.

    Body:
    {
        "rating": 1-5,
        "comment": "optional text"
    }
    """
    try:
        user_id = user.get("user_id") or user.get("id")
        rating = request_body.get("rating")
        comment = request_body.get("comment", "")

        if rating is None or not isinstance(rating, int) or rating < 1 or rating > 5:
            raise HTTPException(status_code=400, detail="rating must be an integer between 1 and 5")

        # Verify video belongs to user
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.get_video_by_id(video_id)
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")
        if str(getattr(video, "user_id", None)) != str(user_id):
            raise HTTPException(status_code=403, detail="Forbidden")

        # Map rating (1-5) to importance_score (0.0-1.0)
        importance_score = (rating - 1) / 4.0  # 1->0.0, 2->0.25, 3->0.5, 4->0.75, 5->1.0

        # Update video_phases with user rating, comment, and importance_score
        # Use try-except for graceful fallback if columns don't exist yet
        try:
            sql_update = text("""
                UPDATE video_phases
                SET user_rating = :rating,
                    user_comment = :comment,
                    importance_score = :importance_score,
                    rated_at = NOW(),
                    updated_at = NOW()
                WHERE video_id = :video_id AND phase_index = :phase_index
            """)
            await db.execute(sql_update, {
                "rating": rating,
                "comment": comment,
                "importance_score": importance_score,
                "video_id": video_id,
                "phase_index": phase_index,
            })
            await db.commit()
        except Exception as db_err:
            await db.rollback()
            # Fallback: try without user_rating/user_comment columns
            try:
                sql_fallback = text("""
                    UPDATE video_phases
                    SET importance_score = :importance_score,
                        updated_at = NOW()
                    WHERE video_id = :video_id AND phase_index = :phase_index
                """)
                await db.execute(sql_fallback, {
                    "importance_score": importance_score,
                    "video_id": video_id,
                    "phase_index": phase_index,
                })
                await db.commit()
            except Exception:
                await db.rollback()
                logger.warning(f"Could not update video_phases for rating: {db_err}")

        # Update Qdrant quality_score for RAG learning (in background for faster response)
        def _update_qdrant_bg(vid, pidx, r, c):
            try:
                from app.services.rag.knowledge_store import update_quality_score_with_comment
                update_quality_score_with_comment(
                    video_id=vid, phase_index=pidx, rating=r, comment=c,
                )
            except ImportError:
                try:
                    from app.services.rag.knowledge_store import update_quality_score
                    old_rating = 1 if r >= 4 else (-1 if r <= 2 else 0)
                    update_quality_score(video_id=vid, phase_index=pidx, rating=old_rating)
                except Exception as rag_err:
                    logger.warning(f"Could not update Qdrant quality_score: {rag_err}")
            except Exception as rag_err:
                logger.warning(f"Could not update Qdrant quality_score: {rag_err}")

        background_tasks.add_task(_update_qdrant_bg, video_id, phase_index, rating, comment)

        logger.info(f"Phase rated: video={video_id}, phase={phase_index}, rating={rating}, comment={comment[:50] if comment else ''}")

        return {
            "success": True,
            "video_id": video_id,
            "phase_index": phase_index,
            "rating": rating,
            "comment": comment,
            "importance_score": importance_score,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to rate phase: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to rate phase: {exc}")


# =========================================================
# Product Exposure Timeline API
# =========================================================

@router.get("/{video_id}/product-exposures")
async def get_product_exposures(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Get AI-detected product exposure timeline for a video.
    Returns list of product exposure segments sorted by time_start.
    """
    try:
        # Verify video belongs to user
        result = await db.execute(
            text("SELECT user_id FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        if row[0] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")

        # Ensure table exists (safe for first-time access)
        try:
            await db.execute(text("""
                CREATE TABLE IF NOT EXISTS video_product_exposures (
                    id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
                    video_id UUID NOT NULL,
                    user_id INTEGER,
                    product_name TEXT NOT NULL,
                    brand_name TEXT,
                    product_image_url TEXT,
                    time_start FLOAT NOT NULL,
                    time_end FLOAT NOT NULL,
                    confidence FLOAT DEFAULT 0.8,
                    source VARCHAR(20) DEFAULT 'ai',
                    created_at TIMESTAMPTZ DEFAULT now(),
                    updated_at TIMESTAMPTZ DEFAULT now()
                )
            """))
            await db.commit()
        except Exception:
            await db.rollback()

        # Fetch exposures
        result = await db.execute(
            text("""
                SELECT id, video_id, user_id, product_name, brand_name,
                       product_image_url, time_start, time_end, confidence, source,
                       created_at, updated_at
                FROM video_product_exposures
                WHERE video_id = :vid
                ORDER BY time_start ASC
            """),
            {"vid": video_id},
        )
        rows = result.fetchall()

        exposures = []
        for r in rows:
            exposures.append({
                "id": str(r[0]),
                "video_id": str(r[1]),
                "user_id": r[2],
                "product_name": r[3],
                "brand_name": r[4],
                "product_image_url": r[5],
                "time_start": r[6],
                "time_end": r[7],
                "confidence": r[8],
                "source": r[9],
                "created_at": r[10].isoformat() if r[10] else None,
                "updated_at": r[11].isoformat() if r[11] else None,
            })

        return {"exposures": exposures, "count": len(exposures)}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to get product exposures: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@router.put("/{video_id}/product-exposures/{exposure_id}")
async def update_product_exposure(
    video_id: str,
    exposure_id: str,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Update a product exposure segment (human edit).
    Payload can include: product_name, brand_name, time_start, time_end, confidence
    """
    try:
        # Verify video belongs to user
        result = await db.execute(
            text("SELECT user_id FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        if row[0] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")

        # Build dynamic SET clause
        allowed_fields = ["product_name", "brand_name", "time_start", "time_end", "confidence"]
        set_parts = []
        params = {"eid": exposure_id, "vid": video_id}

        for field in allowed_fields:
            if field in payload:
                set_parts.append(f"{field} = :{field}")
                params[field] = payload[field]

        if not set_parts:
            raise HTTPException(status_code=400, detail="No fields to update")

        # Mark as human-edited
        set_parts.append("source = 'human'")
        set_parts.append("updated_at = now()")

        sql = text(f"""
            UPDATE video_product_exposures
            SET {', '.join(set_parts)}
            WHERE id = :eid AND video_id = :vid
            RETURNING id
        """)

        result = await db.execute(sql, params)
        updated = result.fetchone()
        await db.commit()

        if not updated:
            raise HTTPException(status_code=404, detail="Exposure not found")

        return {"success": True, "id": str(updated[0])}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to update product exposure: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/{video_id}/product-exposures")
async def create_product_exposure(
    video_id: str,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Manually create a product exposure segment.
    Required: product_name, time_start, time_end
    Optional: brand_name, confidence
    """
    try:
        # Verify video belongs to user
        result = await db.execute(
            text("SELECT user_id FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        if row[0] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")

        product_name = payload.get("product_name")
        time_start = payload.get("time_start")
        time_end = payload.get("time_end")

        if not product_name or time_start is None or time_end is None:
            raise HTTPException(
                status_code=400,
                detail="product_name, time_start, time_end are required",
            )

        sql = text("""
            INSERT INTO video_product_exposures
                (video_id, user_id, product_name, brand_name,
                 time_start, time_end, confidence, source)
            VALUES
                (:vid, :uid, :product_name, :brand_name,
                 :time_start, :time_end, :confidence, 'human')
            RETURNING id
        """)

        result = await db.execute(sql, {
            "vid": video_id,
            "uid": current_user["id"],
            "product_name": product_name,
            "brand_name": payload.get("brand_name", ""),
            "time_start": time_start,
            "time_end": time_end,
            "confidence": payload.get("confidence", 1.0),
        })
        new_row = result.fetchone()
        await db.commit()

        return {"success": True, "id": str(new_row[0])}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to create product exposure: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@router.delete("/{video_id}/product-exposures/{exposure_id}")
async def delete_product_exposure(
    video_id: str,
    exposure_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Delete a product exposure segment."""
    try:
        # Verify video belongs to user
        result = await db.execute(
            text("SELECT user_id FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        if row[0] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")
        result = await db.execute(
            text(""""
                DELETE FROM video_product_exposures
                WHERE id = :eid AND video_id = :vid
                RETURNING id
            """),
            {"eid": exposure_id, "vid": video_id},
        )
        deleted = result.fetchone()
        await db.commit()

        if not deleted:
            raise HTTPException(status_code=404, detail="Exposure not found")

        return {"success": True}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to delete product exposure: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/{video_id}/product-exposures/remap-names")
async def remap_product_exposure_names(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Remap generic product names (Product_0, Product_1, ...) to actual names
    from the Excel product data.
    
    Logic:
    1. Get all exposures for this video
    2. Get the product Excel data (same as product-data endpoint)
    3. Extract unique generic names, sort by index (Product_0, Product_1, ...)
    4. Map each Product_N to the Nth product in the Excel list
    5. Also try to find the actual product_name key in Excel data
    6. Bulk update all exposures with the real product names
    """
    try:
        # Verify video belongs to user
        result = await db.execute(
            text("SELECT user_id, excel_product_blob_url FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        if row[0] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Forbidden")

        product_blob_url = row[1]
        if not product_blob_url:
            return {"success": False, "message": "No product Excel file uploaded for this video", "updated": 0}

        # --- Parse Excel to get product list ---
        import httpx
        import tempfile
        import os as _os
        import openpyxl
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions
        from datetime import timedelta

        # Generate SAS URL
        conn_str = _os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
        account_name = ""
        account_key = ""
        for part in conn_str.split(";"):
            if part.startswith("AccountName="):
                account_name = part.split("=", 1)[1]
            elif part.startswith("AccountKey="):
                account_key = part.split("=", 1)[1]

        from urllib.parse import urlparse, unquote
        parsed = urlparse(product_blob_url)
        path = unquote(parsed.path)
        if path.startswith("/videos/"):
            blob_name = path[len("/videos/"):]
        else:
            blob_name = path.lstrip("/")
            if blob_name.startswith("videos/"):
                blob_name = blob_name[len("videos/"):]

        expiry = datetime.now(timezone.utc) + timedelta(minutes=30)
        sas = generate_blob_sas(
            account_name=account_name,
            container_name="videos",
            blob_name=blob_name,
            account_key=account_key,
            permission=BlobSasPermissions(read=True),
            expiry=expiry,
        )
        sas_url = f"https://{account_name}.blob.core.windows.net/videos/{blob_name}?{sas}"

        # Download and parse Excel
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(sas_url)
            if resp.status_code != 200:
                return {"success": False, "message": f"Failed to download Excel (HTTP {resp.status_code})", "updated": 0}

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(resp.content)
            tmp_path = f.name

        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active
            excel_products = []
            if ws:
                rows_data = list(ws.iter_rows(values_only=True))
                if len(rows_data) >= 2:
                    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                    for data_row in rows_data[1:]:
                        if all(v is None for v in data_row):
                            continue
                        item = {}
                        for i, val in enumerate(data_row):
                            if i < len(headers):
                                item[headers[i]] = val
                        excel_products.append(item)
            wb.close()
        finally:
            _os.unlink(tmp_path)

        if not excel_products:
            return {"success": False, "message": "No products found in Excel file", "updated": 0}

        # --- Build name mapping ---
        # Find the product name column in Excel
        # Try common column names: 商品名, product_name, name, 商品タイトル
        name_keys = ["商品名", "product_name", "name", "商品タイトル", "Name", "Product Name", "商品"]
        product_name_key = None
        sample = excel_products[0]
        for key in name_keys:
            if key in sample and sample[key]:
                product_name_key = key
                break
        # If not found, try first string column
        if not product_name_key:
            for k, v in sample.items():
                if isinstance(v, str) and len(v) > 2:
                    product_name_key = k
                    break

        if not product_name_key:
            return {"success": False, "message": "Could not find product name column in Excel", "updated": 0}

        # Build ordered list of real product names from Excel
        real_names = []
        for p in excel_products:
            pname = p.get(product_name_key)
            if pname:
                real_names.append(str(pname).strip())
            else:
                real_names.append(None)

        logger.info(f"[REMAP] Found {len(real_names)} products in Excel, name_key='{product_name_key}'")
        logger.info(f"[REMAP] First 5 products: {real_names[:5]}")

        # --- Get current exposures ---
        result = await db.execute(
            text("""
                SELECT DISTINCT product_name
                FROM video_product_exposures
                WHERE video_id = :vid
                ORDER BY product_name
            """),
            {"vid": video_id},
        )
        current_names = [r[0] for r in result.fetchall()]

        # Build mapping: Product_N -> real_names[N]
        import re
        name_map = {}
        for cname in current_names:
            match = re.match(r"^Product_(\d+)$", cname)
            if match:
                idx = int(match.group(1))
                if idx < len(real_names) and real_names[idx]:
                    name_map[cname] = real_names[idx]

        if not name_map:
            return {
                "success": False,
                "message": f"No Product_N names found to remap. Current names: {current_names[:10]}",
                "updated": 0,
            }

        logger.info(f"[REMAP] Mapping {len(name_map)} names: {name_map}")

        # --- Bulk update ---
        total_updated = 0
        for old_name, new_name in name_map.items():
            result = await db.execute(
                text("""
                    UPDATE video_product_exposures
                    SET product_name = :new_name, updated_at = now()
                    WHERE video_id = :vid AND product_name = :old_name
                """),
                {"vid": video_id, "old_name": old_name, "new_name": new_name},
            )
            total_updated += result.rowcount

        await db.commit()

        return {
            "success": True,
            "message": f"Remapped {len(name_map)} product names, {total_updated} rows updated",
            "updated": total_updated,
            "mapping": name_map,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to remap product names: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/remap-all-product-names")
async def remap_all_product_names(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Remap product names for ALL videos belonging to the current user.
    Iterates over all videos with product exposures and applies the remap logic.
    """
    try:
        # Get all video IDs for this user that have product exposures
        result = await db.execute(
            text("""
                SELECT DISTINCT vpe.video_id
                FROM video_product_exposures vpe
                JOIN videos v ON vpe.video_id = v.id
                WHERE v.user_id = :uid
                  AND vpe.product_name ~ '^Product_\\d+$'
            """),
            {"uid": current_user["id"]},
        )
        video_ids = [str(r[0]) for r in result.fetchall()]

        if not video_ids:
            return {"success": True, "message": "No videos with generic Product_N names found", "videos_processed": 0}

        results = []
        for vid in video_ids:
            try:
                # Call the single-video remap logic inline
                # (We can't easily call the endpoint from here, so duplicate the core logic)
                vrow = await db.execute(
                    text("SELECT excel_product_blob_url FROM videos WHERE id = :vid"),
                    {"vid": vid},
                )
                vdata = vrow.fetchone()
                if not vdata or not vdata[0]:
                    results.append({"video_id": vid, "status": "skipped", "reason": "no Excel"})
                    continue

                results.append({"video_id": vid, "status": "needs_individual_call"})
            except Exception as e:
                results.append({"video_id": vid, "status": "error", "reason": str(e)})

        return {
            "success": True,
            "message": f"Found {len(video_ids)} videos with generic names. Call /remap-names on each individually.",
            "video_ids": video_ids,
            "details": results,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to list videos for remap: {exc}")
        raise HTTPException(status_code=500, detail=str(exc))


# ============================================================
# TikTok Live Capture Endpoints
# ============================================================

@router.post("/live-check", response_model=LiveCheckResponse)
async def live_check(
    payload: LiveCaptureRequest,
    current_user=Depends(get_current_user),
):
    """Check if a TikTok user is currently live."""
    from app.services.tiktok_service import TikTokLiveService

    try:
        info = await TikTokLiveService.check_and_get_info(payload.live_url)
        return LiveCheckResponse(
            is_live=info["is_live"],
            username=info.get("username"),
            room_id=info.get("room_id"),
            title=info.get("title"),
            message="LIVE" if info["is_live"] else "User is not currently live",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as exc:
        logger.exception(f"Live check failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Live check failed: {exc}")


@router.post("/live-capture", response_model=LiveCaptureResponse)
async def live_capture(
    payload: LiveCaptureRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Start capturing a TikTok live stream.
    1. Validates the URL and checks if the user is live
    2. Creates a video record in the database
    3. Enqueues a live_capture job for the worker
    """
    from app.services.tiktok_service import TikTokLiveService
    from app.services.queue_service import enqueue_job

    # Step 1: Check live status
    try:
        info = await TikTokLiveService.check_and_get_info(payload.live_url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as exc:
        logger.exception(f"Live check failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to check live status: {exc}")

    if not info["is_live"]:
        raise HTTPException(
            status_code=400,
            detail=f"@{info.get('username', 'unknown')} is not currently live",
        )

    username = info["username"]
    title = info.get("title", "")

    # Step 2: Create video record
    video_id = str(uuid_module.uuid4())
    original_filename = f"tiktok_live_{username}.mp4"

    try:
        video_repo = VideoRepository(lambda: db)
        service = VideoService(video_repository=video_repo)

        video = await video_repo.create_video(
            user_id=current_user["id"],
            video_id=video_id,
            original_filename=original_filename,
            status="capturing",
            upload_type="live_capture",
        )
        await db.commit()
    except Exception as exc:
        logger.exception(f"Failed to create video record: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to create video record: {exc}")

    # Step 3: Enqueue live_capture job
    try:
        queue_payload = {
            "job_type": "live_capture",
            "video_id": video_id,
            "live_url": payload.live_url,
            "email": current_user["email"],
            "user_id": current_user["id"],
            "duration": payload.duration or 0,
            "username": username,
            "stream_title": title,
        }
        await enqueue_job(queue_payload)
    except Exception as exc:
        logger.exception(f"Failed to enqueue live capture job: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to start capture: {exc}")

    return LiveCaptureResponse(
        video_id=video_id,
        status="capturing",
        stream_title=title,
        username=username,
        message=f"Live capture started for @{username}; recording and analysis will begin automatically",
    )
