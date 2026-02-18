"""
Excel parser for clean video uploads.
Parses product.xlsx and trend_stats.xlsx files
and returns structured data for report generation.
"""
import os
import logging
import requests
from urllib.parse import urlparse

logger = logging.getLogger("process_video")

AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")


def _parse_conn_str(conn_str: str) -> dict:
    """Parse AccountName and AccountKey from Azure Storage connection string."""
    parts = conn_str.split(";")
    out = {"AccountName": None, "AccountKey": None}
    for p in parts:
        if p.startswith("AccountName="):
            out["AccountName"] = p.split("=", 1)[1]
        if p.startswith("AccountKey="):
            out["AccountKey"] = p.split("=", 1)[1]
    return out


def _ensure_sas_token(blob_url: str) -> str:
    """
    Ensure the blob URL has a SAS token for authentication.
    If no SAS token is present, generate one using the connection string.
    """
    if not blob_url:
        return blob_url

    # Already has SAS token
    if "?" in blob_url and ("sig=" in blob_url or "se=" in blob_url):
        return blob_url

    # No SAS token → generate one
    logger.info("[EXCEL] No SAS token in URL, generating one...")
    try:
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions
        from datetime import datetime, timedelta

        conn = _parse_conn_str(AZURE_STORAGE_CONNECTION_STRING)
        account_name = conn["AccountName"]
        account_key = conn["AccountKey"]

        if not account_name or not account_key:
            logger.warning("[EXCEL] No Azure credentials available for SAS generation")
            return blob_url

        # Parse blob URL to extract container and blob path
        parsed = urlparse(blob_url)
        path_parts = parsed.path.lstrip("/").split("/", 1)
        if len(path_parts) < 2:
            logger.warning("[EXCEL] Cannot parse blob path from URL: %s", blob_url)
            return blob_url

        container_name = path_parts[0]
        blob_name = path_parts[1]

        expiry = datetime.utcnow() + timedelta(minutes=60)
        sas = generate_blob_sas(
            account_name=account_name,
            container_name=container_name,
            blob_name=blob_name,
            account_key=account_key,
            permission=BlobSasPermissions(read=True),
            expiry=expiry,
        )
        url_with_sas = f"{blob_url}?{sas}"
        logger.info("[EXCEL] SAS token generated successfully")
        return url_with_sas

    except Exception as e:
        logger.warning("[EXCEL] Failed to generate SAS token: %s", e)
        return blob_url


def download_excel(blob_url: str, dest_path: str) -> bool:
    """Download an Excel file from Azure Blob URL (with auto SAS token)."""
    if not blob_url:
        return False
    try:
        # Ensure URL has SAS token for authentication
        url = _ensure_sas_token(blob_url)

        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with requests.get(url, stream=True, timeout=60) as r:
            r.raise_for_status()
            with open(dest_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        logger.info(f"[EXCEL] Downloaded: {dest_path}")
        return True
    except Exception as e:
        logger.warning(f"[EXCEL] Download failed: {e}")
        return False


def parse_product_excel(file_path: str) -> list[dict]:
    """
    Parse product.xlsx.
    Expected columns (flexible matching):
    - 商品名 / product_name / name
    - 価格 / price
    - カテゴリ / category
    - その他の列も取り込む
    Returns list of product dicts.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("[EXCEL] openpyxl not installed, cannot parse Excel")
        return []

    if not os.path.exists(file_path):
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        # First row = headers
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        products = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            product = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    product[headers[i]] = val
            products.append(product)

        wb.close()
        logger.info(f"[EXCEL] Parsed {len(products)} products from {file_path}")
        return products

    except Exception as e:
        logger.warning(f"[EXCEL] Failed to parse product Excel: {e}")
        return []


def parse_trend_excel(file_path: str) -> list[dict]:
    """
    Parse trend_stats.xlsx.
    Expected columns (flexible matching):
    - 時間 / time / timestamp
    - 売上 / sales / revenue
    - 注文数 / orders
    - 商品名 / product_name
    - その他の列も取り込む
    Returns list of trend data dicts.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("[EXCEL] openpyxl not installed, cannot parse Excel")
        return []

    if not os.path.exists(file_path):
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        trend_data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            entry = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    entry[headers[i]] = val
            trend_data.append(entry)

        wb.close()
        logger.info(f"[EXCEL] Parsed {len(trend_data)} trend entries from {file_path}")
        return trend_data

    except Exception as e:
        logger.warning(f"[EXCEL] Failed to parse trend Excel: {e}")
        return []


def load_excel_data(video_id: str, excel_urls: dict, work_dir: str = "excel_data") -> dict:
    """
    Download and parse both Excel files for a video.
    Returns dict with 'products' and 'trends' keys.
    """
    excel_dir = os.path.join(work_dir, video_id)
    os.makedirs(excel_dir, exist_ok=True)

    result = {
        "products": [],
        "trends": [],
        "has_product_data": False,
        "has_trend_data": False,
    }

    # Download and parse product data
    product_url = excel_urls.get("excel_product_blob_url")
    if product_url:
        product_path = os.path.join(excel_dir, "product.xlsx")
        if download_excel(product_url, product_path):
            result["products"] = parse_product_excel(product_path)
            result["has_product_data"] = len(result["products"]) > 0

    # Download and parse trend data
    trend_url = excel_urls.get("excel_trend_blob_url")
    if trend_url:
        trend_path = os.path.join(excel_dir, "trend_stats.xlsx")
        if download_excel(trend_url, trend_path):
            result["trends"] = parse_trend_excel(trend_path)
            result["has_trend_data"] = len(result["trends"]) > 0

    logger.info(
        f"[EXCEL] Loaded data for {video_id}: "
        f"{len(result['products'])} products, "
        f"{len(result['trends'])} trend entries"
    )

    return result


def format_excel_data_for_prompt(excel_data: dict) -> str:
    """
    Format Excel data into a text summary for GPT prompts.
    """
    parts = []

    if excel_data.get("has_product_data"):
        parts.append("【商品データ】")
        for i, p in enumerate(excel_data["products"][:20], 1):  # Max 20 products
            items = [f"{k}: {v}" for k, v in p.items() if v is not None]
            parts.append(f"  {i}. " + " / ".join(items))

    if excel_data.get("has_trend_data"):
        parts.append("\n【売上トレンドデータ】")
        for i, t in enumerate(excel_data["trends"][:50], 1):  # Max 50 entries
            items = [f"{k}: {v}" for k, v in t.items() if v is not None]
            parts.append(f"  {i}. " + " / ".join(items))

    return "\n".join(parts) if parts else ""


def match_sales_to_phase(trends: list[dict], start_sec: float, end_sec: float) -> dict:
    """
    Match trend/sales data to a specific phase time range.
    Returns aggregated sales metrics for the phase.

    Tries to match using time-based columns.
    """
    import re
    from datetime import datetime, timedelta

    if not trends:
        return {"sales": None, "orders": None, "products_sold": []}

    # Detect time column
    time_keys = []
    sales_keys = []
    order_keys = []
    product_keys = []

    sample = trends[0]
    for k in sample.keys():
        kl = k.lower()
        if any(w in kl for w in ["時間", "time", "timestamp", "秒", "sec", "minute"]):
            time_keys.append(k)
        if any(w in kl for w in ["売上", "sales", "revenue", "金額", "amount"]):
            sales_keys.append(k)
        if any(w in kl for w in ["注文", "order", "件数", "count"]):
            order_keys.append(k)
        if any(w in kl for w in ["商品", "product", "item", "名前", "name"]):
            product_keys.append(k)

    phase_sales = 0
    phase_orders = 0
    products_sold = []

    for t in trends:
        # Try to extract time in seconds from the entry
        entry_time = None
        for tk in time_keys:
            val = t.get(tk)
            if val is None:
                continue
            try:
                # Try direct seconds
                entry_time = float(val)
                break
            except (ValueError, TypeError):
                pass
            try:
                # Try MM:SS or HH:MM:SS format
                val_str = str(val)
                parts = val_str.split(":")
                if len(parts) == 2:
                    entry_time = int(parts[0]) * 60 + int(parts[1])
                    break
                elif len(parts) == 3:
                    entry_time = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    break
            except (ValueError, TypeError):
                pass

        if entry_time is None:
            continue

        # Check if this entry falls within the phase
        if start_sec <= entry_time <= end_sec:
            for sk in sales_keys:
                try:
                    phase_sales += float(t.get(sk, 0) or 0)
                except (ValueError, TypeError):
                    pass
            for ok in order_keys:
                try:
                    phase_orders += int(t.get(ok, 0) or 0)
                except (ValueError, TypeError):
                    pass
            for pk in product_keys:
                pname = t.get(pk)
                if pname and str(pname).strip():
                    products_sold.append(str(pname).strip())

    return {
        "sales": phase_sales if phase_sales > 0 else None,
        "orders": phase_orders if phase_orders > 0 else None,
        "products_sold": list(set(products_sold)),
    }


def build_phase_stats_from_csv(
    trends: list[dict],
    keyframes: list[int],
    total_frames: int,
    video_start_time_sec: float | None = None,
) -> list[dict]:
    """
    STEP 2 代替 – CSVトレンドデータからphase_statsを生成。
    GPT Vision APIを一切呼ばずに、CSVの数値をそのまま使う。

    クリーン動画の場合、TikTokのUIが映っていないため
    画面からのOCR読み取りは不要。CSVに全データがある。

    Returns:
        extract_phase_stats() と同じ形式のリスト:
        [{
            "phase_index": int,
            "phase_start_frame": int,
            "phase_end_frame": int,
            "phase_start_used_frame": int,
            "start": {"viewer_count": int|None, "like_count": int|None},
            "phase_end_used_frame": int,
            "end": {"viewer_count": int|None, "like_count": int|None},
        }]
    """
    from csv_slot_filter import _parse_time_to_seconds, _detect_time_key, _find_key

    extended = [0] + keyframes + [total_frames]

    # CSVの時刻カラムを検出
    time_key = _detect_time_key(trends)
    if not time_key:
        logger.warning("[CSV_STATS] No time column found, returning empty stats")
        return _build_empty_stats(extended)

    # CSVの各エントリを秒数に変換してソート
    timed_entries = []
    for entry in trends:
        t_sec = _parse_time_to_seconds(entry.get(time_key))
        if t_sec is not None:
            timed_entries.append({"time_sec": t_sec, "entry": entry})
    timed_entries.sort(key=lambda x: x["time_sec"])

    if not timed_entries:
        logger.warning("[CSV_STATS] No valid time entries found")
        return _build_empty_stats(extended)

    # 動画開始時刻の推定
    if video_start_time_sec is None:
        video_start_time_sec = timed_entries[0]["time_sec"]

    # viewer_count / like_count のカラム名を検出
    sample = trends[0]
    viewer_key = _find_key(sample, ["观看人数", "viewers", "viewer_count", "観看人数"])
    like_key = _find_key(sample, ["点赞数", "likes", "like_count", "いいね数"])

    logger.info(
        "[CSV_STATS] Building phase stats from CSV: "
        "time_key=%s, viewer_key=%s, like_key=%s, "
        "video_start=%d sec, %d entries",
        time_key, viewer_key, like_key,
        video_start_time_sec, len(timed_entries),
    )

    results = []
    for i in range(len(extended) - 1):
        phase_start_frame = extended[i]
        phase_end_frame = extended[i + 1] - 1
        phase_idx = i + 1

        # フレーム番号を動画内秒数に変換（fps=1前提）
        phase_start_sec = float(phase_start_frame)
        phase_end_sec = float(phase_end_frame)

        # フェーズの開始/終了に最も近いCSVエントリを見つける
        start_metrics = _find_nearest_csv_metrics(
            timed_entries, phase_start_sec + video_start_time_sec,
            viewer_key, like_key,
        )
        end_metrics = _find_nearest_csv_metrics(
            timed_entries, phase_end_sec + video_start_time_sec,
            viewer_key, like_key,
        )

        results.append({
            "phase_index": phase_idx,
            "phase_start_frame": phase_start_frame,
            "phase_start_used_frame": phase_start_frame,
            "start": start_metrics,
            "phase_end_frame": phase_end_frame,
            "phase_end_used_frame": phase_end_frame,
            "end": end_metrics,
        })

    logger.info("[CSV_STATS] Built stats for %d phases (0 API calls)", len(results))
    return results


def _find_nearest_csv_metrics(
    timed_entries: list[dict],
    target_sec: float,
    viewer_key: str | None,
    like_key: str | None,
) -> dict | None:
    """
    target_secに最も近いCSVエントリからviewer_count/like_countを取得。
    """
    if not timed_entries:
        return None

    best_entry = None
    best_diff = float("inf")

    for te in timed_entries:
        diff = abs(te["time_sec"] - target_sec)
        if diff < best_diff:
            best_diff = diff
            best_entry = te["entry"]

    if best_entry is None:
        return None

    viewer_count = None
    like_count = None

    if viewer_key:
        try:
            val = best_entry.get(viewer_key)
            if val is not None:
                viewer_count = int(float(val))
        except (ValueError, TypeError):
            pass

    if like_key:
        try:
            val = best_entry.get(like_key)
            if val is not None:
                like_count = int(float(val))
        except (ValueError, TypeError):
            pass

    if viewer_count is not None or like_count is not None:
        return {"viewer_count": viewer_count, "like_count": like_count}

    return None


def _build_empty_stats(extended: list[int]) -> list[dict]:
    """CSVデータが不足している場合の空のphase_statsを生成"""
    results = []
    for i in range(len(extended) - 1):
        results.append({
            "phase_index": i + 1,
            "phase_start_frame": extended[i],
            "phase_start_used_frame": extended[i],
            "start": None,
            "phase_end_frame": extended[i + 1] - 1,
            "phase_end_used_frame": extended[i + 1] - 1,
            "end": None,
        })
    return results
